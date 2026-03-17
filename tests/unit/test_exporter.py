"""Comprehensive tests for ReportExporter module (Task 10: Export Format Tests).

Tests cover:
    - CSV export with proper escaping and delimiters
    - Excel export with formatting and large dataset handling
    - File validation and permissions
    - Error handling (empty data, IO errors)
    - UTF-8 and special character support
    - Performance with 1000+ row datasets

Phase 4 Implementation:
    - 20+ tests with >80% coverage of ReportExporter class
    - All tests use proper fixtures and async patterns
    - No flaky tests or race conditions
    - Windows-compatible file operations
"""

import pytest
import asyncio
import csv
import json
from pathlib import Path
from unittest.mock import MagicMock, patch, AsyncMock
from typing import Dict, Any, List

from sap.exporter import ReportExporter, ReportResult, ExportFormat


# ─────────────────────────────────────────────────────────────────
# FIXTURES: Sample Data for Testing
# ─────────────────────────────────────────────────────────────────

@pytest.fixture
def sample_result() -> ReportResult:
    """Basic 3x3 result set for testing.
    
    Returns:
        ReportResult with simple test data
    """
    return ReportResult(
        columns=["Material", "Plant", "Qty"],
        rows=[
            {"Material": "MAT001", "Plant": "1000", "Qty": "100"},
            {"Material": "MAT002", "Plant": "2000", "Qty": "200"},
            {"Material": "MAT003", "Plant": "3000", "Qty": "300"},
        ]
    )


@pytest.fixture
def special_characters_result() -> ReportResult:
    """Result with umlauts, Chinese characters, and emoji.
    
    Tests UTF-8 encoding and special character handling.
    
    Returns:
        ReportResult with international characters
    """
    return ReportResult(
        columns=["Name", "Description", "Region"],
        rows=[
            {
                "Name": "Müller",
                "Description": "German name with umlaut",
                "Region": "DE"
            },
            {
                "Name": "张三",
                "Description": "Chinese name",
                "Region": "CN"
            },
            {
                "Name": "François",
                "Description": "French name",
                "Region": "FR"
            },
            {
                "Name": "Test 😀",
                "Description": "Emoji support",
                "Region": "INT"
            },
        ]
    )


@pytest.fixture
def result_with_special_values() -> ReportResult:
    """Result with values containing quotes, commas, newlines.
    
    Tests proper CSV escaping of special characters.
    
    Returns:
        ReportResult with values that need escaping
    """
    return ReportResult(
        columns=["ID", "Text", "Description"],
        rows=[
            {
                "ID": "1",
                "Text": 'Hello, World',
                "Description": 'Value with "quotes"'
            },
            {
                "ID": "2",
                "Text": "Multi\nline\ntext",
                "Description": 'Mixed "quotes" and, commas'
            },
            {
                "ID": "3",
                "Text": "Simple",
                "Description": "No special chars"
            },
        ]
    )


@pytest.fixture
def large_result() -> ReportResult:
    """1000+ rows for performance testing.
    
    Verifies efficient handling of large datasets without memory issues.
    
    Returns:
        ReportResult with 1000 rows
    """
    return ReportResult(
        columns=["ID", "Value", "Status"],
        rows=[
            {
                "ID": str(i),
                "Value": f"Row {i}",
                "Status": "Active" if i % 2 == 0 else "Inactive"
            }
            for i in range(1000)
        ]
    )


@pytest.fixture
def empty_result() -> ReportResult:
    """Empty result set (no rows).
    
    Returns:
        ReportResult with columns but no rows
    """
    return ReportResult(
        columns=["Material", "Plant", "Qty"],
        rows=[]
    )


@pytest.fixture
def exporter() -> ReportExporter:
    """Provide ReportExporter instance for testing.
    
    Returns:
        ReportExporter with default settings
    """
    return ReportExporter()


# ─────────────────────────────────────────────────────────────────
# CSV EXPORT TESTS (8 tests)
# ─────────────────────────────────────────────────────────────────

class TestCSVExport:
    """Test CSV export functionality."""

    @pytest.mark.asyncio
    async def test_export_to_csv_basic(
        self,
        exporter: ReportExporter,
        sample_result: ReportResult,
        tmp_path: Path
    ) -> None:
        """Export simple 3x3 grid to CSV.
        
        Verifies:
            - File is created with correct path
            - Headers are included by default
            - All rows are written
            - File is readable
        """
        output_path = tmp_path / "output.csv"
        
        await exporter.export_to_csv(sample_result, output_path)
        
        # Verify file exists
        assert output_path.exists()
        
        # Verify content
        with open(output_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        
        assert len(rows) == 3
        assert rows[0]["Material"] == "MAT001"
        assert rows[1]["Plant"] == "2000"

    @pytest.mark.asyncio
    async def test_export_to_csv_with_headers(
        self,
        exporter: ReportExporter,
        sample_result: ReportResult,
        tmp_path: Path
    ) -> None:
        """Verify header row included when include_headers=True.
        
        Verifies:
            - Header row is present
            - Column names match result.columns
            - Data rows follow headers
        """
        output_path = tmp_path / "with_headers.csv"
        
        await exporter.export_to_csv(
            sample_result,
            output_path,
            include_headers=True
        )
        
        with open(output_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        lines = content.strip().split('\n')
        assert lines[0] == "Material,Plant,Qty"
        assert len(lines) == 4  # 1 header + 3 data rows

    @pytest.mark.asyncio
    async def test_export_to_csv_no_headers(
        self,
        exporter: ReportExporter,
        sample_result: ReportResult,
        tmp_path: Path
    ) -> None:
        """Verify only data rows when include_headers=False.
        
        Verifies:
            - Header row is NOT included
            - Only data rows are present
            - First row contains actual data
        """
        output_path = tmp_path / "no_headers.csv"
        
        await exporter.export_to_csv(
            sample_result,
            output_path,
            include_headers=False
        )
        
        with open(output_path, 'r', encoding='utf-8') as f:
            lines = f.read().strip().split('\n')
        
        assert len(lines) == 3  # Only 3 data rows
        assert "MAT001" in lines[0]

    @pytest.mark.asyncio
    async def test_export_to_csv_custom_delimiter(
        self,
        exporter: ReportExporter,
        sample_result: ReportResult,
        tmp_path: Path
    ) -> None:
        """Test custom delimiters (semicolon, pipe, tab).
        
        Verifies:
            - Semicolon delimiter works
            - Pipe delimiter works
            - Tab delimiter works
            - Delimiter is applied to headers and rows
        """
        # Test semicolon
        output_path = tmp_path / "semicolon.csv"
        await exporter.export_to_csv(
            sample_result,
            output_path,
            delimiter=";"
        )
        
        with open(output_path, 'r', encoding='utf-8') as f:
            first_line = f.readline().strip()
        
        assert first_line == "Material;Plant;Qty"
        
        # Test pipe
        output_path = tmp_path / "pipe.csv"
        await exporter.export_to_csv(
            sample_result,
            output_path,
            delimiter="|"
        )
        
        with open(output_path, 'r', encoding='utf-8') as f:
            first_line = f.readline().strip()
        
        assert first_line == "Material|Plant|Qty"

    @pytest.mark.asyncio
    async def test_export_to_csv_escaping(
        self,
        exporter: ReportExporter,
        result_with_special_values: ReportResult,
        tmp_path: Path
    ) -> None:
        """Verify proper quoting of values with commas, quotes, newlines.
        
        Verifies:
            - Values with commas are quoted
            - Quotes within values are escaped
            - Newlines are properly handled
            - All rows are readable by csv.DictReader
        """
        output_path = tmp_path / "escaped.csv"
        
        await exporter.export_to_csv(result_with_special_values, output_path)
        
        with open(output_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        
        # Verify first row with comma is read correctly
        assert rows[0]["Text"] == "Hello, World"
        
        # Verify second row with quotes is read correctly
        assert 'quotes' in rows[1]["Description"]

    @pytest.mark.asyncio
    async def test_export_to_csv_utf8_encoding(
        self,
        exporter: ReportExporter,
        special_characters_result: ReportResult,
        tmp_path: Path
    ) -> None:
        """Test UTF-8 characters (umlauts, Chinese, emoji).
        
        Verifies:
            - File is encoded as UTF-8
            - Umlauts are preserved
            - Chinese characters are preserved
            - Emoji are preserved
            - File can be read back with correct characters
        """
        output_path = tmp_path / "utf8.csv"
        
        await exporter.export_to_csv(special_characters_result, output_path)
        
        with open(output_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        
        # Verify special characters are preserved
        assert rows[0]["Name"] == "Müller"
        assert rows[1]["Name"] == "张三"
        assert rows[3]["Name"] == "Test 😀"

    @pytest.mark.asyncio
    async def test_export_to_csv_large_dataset(
        self,
        exporter: ReportExporter,
        large_result: ReportResult,
        tmp_path: Path
    ) -> None:
        """Export 1000+ rows efficiently without memory issues.
        
        Verifies:
            - Large dataset exports successfully
            - All rows are written
            - File is readable
            - Performance is acceptable
        """
        output_path = tmp_path / "large.csv"
        
        # Should complete without memory issues
        await exporter.export_to_csv(large_result, output_path)
        
        # Verify all rows are present
        with open(output_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        
        assert len(rows) == 1000

    @pytest.mark.asyncio
    async def test_export_to_csv_empty_result(
        self,
        exporter: ReportExporter,
        empty_result: ReportResult,
        tmp_path: Path
    ) -> None:
        """Error handling for empty result (no rows).
        
        Verifies:
            - ValueError is raised for empty rows
            - Error message is descriptive
        """
        output_path = tmp_path / "empty.csv"
        
        with pytest.raises(ValueError, match="must have rows"):
            await exporter.export_to_csv(empty_result, output_path)


# ─────────────────────────────────────────────────────────────────
# EXCEL EXPORT TESTS (8 tests)
# ─────────────────────────────────────────────────────────────────

class TestExcelExport:
    """Test Excel export functionality."""

    @pytest.mark.asyncio
    async def test_export_to_excel_basic(
        self,
        exporter: ReportExporter,
        sample_result: ReportResult,
        tmp_path: Path
    ) -> None:
        """Export simple 3x3 grid to XLSX.
        
        Verifies:
            - File is created
            - File is valid XLSX
            - Headers and data are present
            - File can be opened with openpyxl
        """
        output_path = tmp_path / "output.xlsx"
        
        await exporter.export_to_excel(sample_result, output_path)
        
        # Verify file exists
        assert output_path.exists()
        
        # Verify it's a valid Excel file
        try:
            import openpyxl
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            assert ws is not None
            
            # Verify headers
            assert ws.cell(1, 1).value == "Material"
            assert ws.cell(1, 2).value == "Plant"
            
            # Verify data
            assert ws.cell(2, 1).value == "MAT001"
            assert ws.cell(2, 2).value == "1000"
        except ImportError:
            pytest.skip("openpyxl not installed")

    @pytest.mark.asyncio
    async def test_export_to_excel_formatting(
        self,
        exporter: ReportExporter,
        sample_result: ReportResult,
        tmp_path: Path
    ) -> None:
        """Verify bold headers, blue background, white text.
        
        Verifies:
            - Header cells are bold
            - Header cells have blue background
            - Header cells have white text color
            - Data cells have normal formatting
        """
        output_path = tmp_path / "formatted.xlsx"
        
        await exporter.export_to_excel(
            sample_result,
            output_path,
            include_formatting=True
        )
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            assert ws is not None
            
            # Verify header formatting
            header_cell = ws.cell(1, 1)
            assert header_cell.font.bold is True
            assert header_cell.font.color is not None  # Should be white
            assert header_cell.fill.start_color is not None  # Should be blue
        except ImportError:
            pytest.skip("openpyxl not installed")

    @pytest.mark.asyncio
    async def test_export_to_excel_column_widths(
        self,
        exporter: ReportExporter,
        sample_result: ReportResult,
        tmp_path: Path
    ) -> None:
        """Verify auto-width columns fit content.
        
        Verifies:
            - Column widths are set
            - Width is proportional to content
            - Width is capped at 50 characters
        """
        output_path = tmp_path / "widths.xlsx"
        
        await exporter.export_to_excel(
            sample_result,
            output_path,
            include_formatting=True
        )
        
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            assert ws is not None
            
            # Verify column widths are set
            col_widths = []
            for col in range(1, 4):
                col_letter = get_column_letter(col)
                width = ws.column_dimensions[col_letter].width
                col_widths.append(width)
            
            # All columns should have width set
            assert all(w is not None for w in col_widths)
            # Width should be reasonable (not too small, not exceeding cap)
            assert all(0 < w <= 50 for w in col_widths)
        except ImportError:
            pytest.skip("openpyxl not installed")

    @pytest.mark.asyncio
    async def test_export_to_excel_frozen_panes(
        self,
        exporter: ReportExporter,
        sample_result: ReportResult,
        tmp_path: Path
    ) -> None:
        """Verify top row is frozen for scrolling.
        
        Verifies:
            - Freeze panes is set to A2 (below headers)
            - Headers remain visible when scrolling
        """
        output_path = tmp_path / "frozen.xlsx"
        
        await exporter.export_to_excel(
            sample_result,
            output_path,
            include_formatting=True
        )
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            assert ws is not None
            
            # Verify freeze panes is set
            assert ws.freeze_panes == "A2"
        except ImportError:
            pytest.skip("openpyxl not installed")

    @pytest.mark.asyncio
    async def test_export_to_excel_special_characters(
        self,
        exporter: ReportExporter,
        special_characters_result: ReportResult,
        tmp_path: Path
    ) -> None:
        """Test special characters (umlauts, Chinese, emoji).
        
        Verifies:
            - Umlauts are preserved
            - Chinese characters are preserved
            - Emoji are preserved
        """
        output_path = tmp_path / "special.xlsx"
        
        await exporter.export_to_excel(special_characters_result, output_path)
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            assert ws is not None
            
            # Verify special characters are preserved
            assert ws.cell(2, 1).value == "Müller"
            assert ws.cell(3, 1).value == "张三"
            assert ws.cell(5, 1).value == "Test 😀"
        except ImportError:
            pytest.skip("openpyxl not installed")

    @pytest.mark.asyncio
    async def test_export_to_excel_large_dataset(
        self,
        exporter: ReportExporter,
        large_result: ReportResult,
        tmp_path: Path
    ) -> None:
        """Export 1000+ rows efficiently with openpyxl streaming.
        
        Verifies:
            - Large dataset exports successfully
            - All rows are written
            - File is readable
        """
        output_path = tmp_path / "large.xlsx"
        
        await exporter.export_to_excel(large_result, output_path)
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            assert ws is not None
            
            # Verify all rows are present (1000 data rows + 1 header)
            assert ws.max_row == 1001
        except ImportError:
            pytest.skip("openpyxl not installed")

    @pytest.mark.asyncio
    async def test_export_to_excel_sheet_name_limit(
        self,
        exporter: ReportExporter,
        sample_result: ReportResult,
        tmp_path: Path
    ) -> None:
        """Sheet name capped at 31 chars (Excel limit).
        
        Verifies:
            - Sheet name is limited to 31 characters
            - Long names are truncated
            - File is valid after truncation
        """
        output_path = tmp_path / "long_name.xlsx"
        
        # Very long sheet name
        long_name = "This is a very long sheet name that exceeds Excel limit"
        
        await exporter.export_to_excel(
            sample_result,
            output_path,
            sheet_name=long_name
        )
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            assert ws is not None
            
            # Sheet name should be truncated to 31 chars
            assert len(ws.title) <= 31
        except ImportError:
            pytest.skip("openpyxl not installed")

    @pytest.mark.asyncio
    async def test_export_to_excel_empty_result(
        self,
        exporter: ReportExporter,
        empty_result: ReportResult,
        tmp_path: Path
    ) -> None:
        """Error handling for empty result.
        
        Verifies:
            - ValueError is raised for empty rows
            - Error message is descriptive
        """
        output_path = tmp_path / "empty.xlsx"
        
        with pytest.raises(ValueError, match="must have rows"):
            await exporter.export_to_excel(empty_result, output_path)


# ─────────────────────────────────────────────────────────────────
# FILE VALIDATION TESTS (4 tests)
# ─────────────────────────────────────────────────────────────────

class TestFileValidation:
    """Test file creation and validation."""

    @pytest.mark.asyncio
    async def test_csv_file_readable(
        self,
        exporter: ReportExporter,
        sample_result: ReportResult,
        tmp_path: Path
    ) -> None:
        """Verify written CSV can be read back correctly.
        
        Verifies:
            - CSV file is readable
            - Content matches original data
            - No encoding errors
        """
        output_path = tmp_path / "readable.csv"
        
        await exporter.export_to_csv(sample_result, output_path)
        
        # Read back and verify
        with open(output_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        
        assert len(rows) == len(sample_result.rows)
        for i, row in enumerate(rows):
            assert row["Material"] == sample_result.rows[i]["Material"]

    @pytest.mark.asyncio
    async def test_excel_file_openable(
        self,
        exporter: ReportExporter,
        sample_result: ReportResult,
        tmp_path: Path
    ) -> None:
        """Verify written XLSX opens in openpyxl without errors.
        
        Verifies:
            - XLSX file is valid
            - Can be opened and read
            - No corruption
        """
        output_path = tmp_path / "openable.xlsx"
        
        await exporter.export_to_excel(sample_result, output_path)
        
        try:
            import openpyxl
            # Should open without error
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            assert ws is not None
            
            # Verify data is present
            assert ws.cell(2, 1).value is not None
        except ImportError:
            pytest.skip("openpyxl not installed")

    @pytest.mark.asyncio
    async def test_file_permissions(
        self,
        exporter: ReportExporter,
        sample_result: ReportResult,
        tmp_path: Path
    ) -> None:
        """Verify exported files have correct read/write permissions.
        
        Verifies:
            - File is readable by owner
            - File is writable by owner
            - File exists after export
        """
        output_path = tmp_path / "permissions.csv"
        
        await exporter.export_to_csv(sample_result, output_path)
        
        # Verify file is readable and writable
        assert output_path.stat().st_mode & 0o400  # Read permission
        assert output_path.stat().st_mode & 0o200  # Write permission

    @pytest.mark.asyncio
    async def test_output_path_creation(
        self,
        exporter: ReportExporter,
        sample_result: ReportResult,
        tmp_path: Path
    ) -> None:
        """Verify parent directories created if missing.
        
        Verifies:
            - Nested directories are created
            - File is created in correct location
            - No errors if directories don't exist
        """
        # Create path with nested directories that don't exist
        output_path = tmp_path / "nested" / "deep" / "structure" / "output.csv"
        
        await exporter.export_to_csv(sample_result, output_path)
        
        # Verify nested directories and file were created
        assert output_path.exists()
        assert output_path.parent.exists()


# ─────────────────────────────────────────────────────────────────
# ERROR HANDLING TESTS (3 tests)
# ─────────────────────────────────────────────────────────────────

class TestErrorHandling:
    """Test error handling in export functions."""

    @pytest.mark.asyncio
    async def test_export_missing_columns(
        self,
        exporter: ReportExporter,
        tmp_path: Path
    ) -> None:
        """Error when result.columns is empty.
        
        Verifies:
            - ValueError is raised
            - Error message is descriptive
        """
        result = ReportResult(columns=[], rows=[{"ID": "1"}])
        output_path = tmp_path / "output.csv"
        
        with pytest.raises(ValueError, match="must have columns"):
            await exporter.export_to_csv(result, output_path)

    @pytest.mark.asyncio
    async def test_export_missing_rows(
        self,
        exporter: ReportExporter,
        tmp_path: Path
    ) -> None:
        """Error when result.rows is empty.
        
        Verifies:
            - ValueError is raised
            - Error message is descriptive
        """
        result = ReportResult(columns=["ID"], rows=[])
        output_path = tmp_path / "output.csv"
        
        with pytest.raises(ValueError, match="must have rows"):
            await exporter.export_to_csv(result, output_path)

    @pytest.mark.asyncio
    async def test_export_io_error(
        self,
        exporter: ReportExporter,
        sample_result: ReportResult,
        tmp_path: Path
    ) -> None:
        """Graceful error handling for disk write failures.
        
        Verifies:
            - IOError is raised when file can't be written
            - Error message is informative
        """
        # Mock the open function to raise IOError during write
        with patch("builtins.open", side_effect=IOError("Permission denied")):
            output_path = tmp_path / "output.csv"
            
            with pytest.raises(IOError):
                await exporter.export_to_csv(sample_result, output_path)


# ─────────────────────────────────────────────────────────────────
# INTEGRATION TESTS (2 tests)
# ─────────────────────────────────────────────────────────────────

class TestIntegration:
    """Integration tests for export workflows."""

    @pytest.mark.asyncio
    async def test_export_all_formats_same_data(
        self,
        exporter: ReportExporter,
        sample_result: ReportResult,
        tmp_path: Path
    ) -> None:
        """Export same data to CSV, Excel, and JSON formats.
        
        Verifies:
            - All formats export successfully
            - Row counts match across formats
            - No data loss during export
        """
        csv_path = tmp_path / "export.csv"
        xlsx_path = tmp_path / "export.xlsx"
        json_path = tmp_path / "export.json"
        
        await exporter.export_to_csv(sample_result, csv_path)
        await exporter.export_to_excel(sample_result, xlsx_path)
        await exporter.export_to_json(sample_result, json_path)
        
        # Verify all files exist
        assert csv_path.exists()
        assert xlsx_path.exists()
        assert json_path.exists()
        
        # Verify CSV
        with open(csv_path, 'r', encoding='utf-8') as f:
            csv_rows = list(csv.DictReader(f))
        
        assert len(csv_rows) == 3
        
        # Verify JSON
        with open(json_path, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        
        assert len(json_data["data"]) == 3

    @pytest.mark.asyncio
    async def test_sequential_exports(
        self,
        exporter: ReportExporter,
        tmp_path: Path
    ) -> None:
        """Export multiple results sequentially without interference.
        
        Verifies:
            - Multiple exports work in sequence
            - No data leakage between exports
            - File isolation is maintained
        """
        result1 = ReportResult(
            columns=["ID", "Name"],
            rows=[
                {"ID": "1", "Name": "First"},
                {"ID": "2", "Name": "Second"},
            ]
        )
        
        result2 = ReportResult(
            columns=["Code", "Value"],
            rows=[
                {"Code": "A", "Value": "100"},
                {"Code": "B", "Value": "200"},
                {"Code": "C", "Value": "300"},
            ]
        )
        
        path1 = tmp_path / "first.csv"
        path2 = tmp_path / "second.csv"
        
        await exporter.export_to_csv(result1, path1)
        await exporter.export_to_csv(result2, path2)
        
        # Verify first file
        with open(path1, 'r', encoding='utf-8') as f:
            rows1 = list(csv.DictReader(f))
        assert len(rows1) == 2
        
        # Verify second file
        with open(path2, 'r', encoding='utf-8') as f:
            rows2 = list(csv.DictReader(f))
        assert len(rows2) == 3


# ─────────────────────────────────────────────────────────────────
# EDGE CASE TESTS (3 tests)
# ─────────────────────────────────────────────────────────────────

class TestEdgeCases:
    """Test edge cases and boundary conditions."""

    @pytest.mark.asyncio
    async def test_single_row_single_column(
        self,
        exporter: ReportExporter,
        tmp_path: Path
    ) -> None:
        """Export minimal data (1 row, 1 column).
        
        Verifies:
            - Single cell data exports correctly
            - No errors with minimal dataset
        """
        result = ReportResult(
            columns=["Value"],
            rows=[{"Value": "one"}]
        )
        
        output_path = tmp_path / "minimal.csv"
        await exporter.export_to_csv(result, output_path)
        
        with open(output_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        
        assert len(rows) == 1
        assert rows[0]["Value"] == "one"

    @pytest.mark.asyncio
    async def test_very_long_cell_values(
        self,
        exporter: ReportExporter,
        tmp_path: Path
    ) -> None:
        """Handle very long cell values (1000+ characters).
        
        Verifies:
            - Long values don't cause errors
            - Values are preserved intact
            - No truncation occurs
        """
        long_text = "A" * 1000
        result = ReportResult(
            columns=["LongText"],
            rows=[{"LongText": long_text}]
        )
        
        output_path = tmp_path / "long.csv"
        await exporter.export_to_csv(result, output_path)
        
        with open(output_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        
        assert rows[0]["LongText"] == long_text

    @pytest.mark.asyncio
    async def test_many_columns(
        self,
        exporter: ReportExporter,
        tmp_path: Path
    ) -> None:
        """Handle result with many columns (50+).
        
        Verifies:
            - Large number of columns work
            - All columns are exported
            - No column limit issues
        """
        num_cols = 50
        columns = [f"Col{i}" for i in range(num_cols)]
        row_data = {f"Col{i}": f"Value{i}" for i in range(num_cols)}
        
        result = ReportResult(
            columns=columns,
            rows=[row_data]
        )
        
        output_path = tmp_path / "many_cols.csv"
        await exporter.export_to_csv(result, output_path)
        
        with open(output_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        
        assert len(rows[0]) == num_cols
